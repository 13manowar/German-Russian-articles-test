import sys
import random
import openpyxl

from PyQt6 import uic
from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox, QDialog

EXCEL_FILE = "Книга 1.xlsx"


class App(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("main.ui", self)

        self.wb = openpyxl.load_workbook(EXCEL_FILE)

        self.all_words = []
        self.filtered_words = []
        self.current = None
        self.session_errors = 0

        # buttons
        self.btnDer.clicked.connect(lambda: self.check("der"))
        self.btnDie.clicked.connect(lambda: self.check("die"))
        self.btnDas.clicked.connect(lambda: self.check("das"))

        self.btnEdit.clicked.connect(self.open_editor)  # 🔥 теперь редактор

        # checkboxes levels
        self.chkLevels = {
            "A1": self.chkA1,
            "A2": self.chkA2,
            "B1": self.chkB1,
            "B2": self.chkB2,
            "C1+": self.chkC1
        }

        for cb in self.chkLevels.values():
            cb.stateChanged.connect(self.reload_words)

        self.chkShowArticle.stateChanged.connect(self.update_view)
        self.chkNoColor.stateChanged.connect(self.update_view)
        self.chkHideTranslation.stateChanged.connect(self.update_view)

        self.btnReset.clicked.connect(self.reset_errors)
        self.btnAdd.clicked.connect(self.add_word)

        self.load_words()
        self.reload_words()

    # ---------------- LOAD ----------------
    def load_words(self):
        self.all_words.clear()

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            for row in range(2, ws.max_row + 1):
                word = ws[f"A{row}"].value
                if not word:
                    continue

                self.all_words.append({
                    "sheet": ws,
                    "row": row,
                    "word": str(word),
                    "translation": ws[f"B{row}"].value or "",
                    "level": str(ws[f"C{row}"].value or "").strip(),
                    "skip": ws[f"E{row}"].value
                })

    # ---------------- FILTER ----------------
    def reload_words(self):
        selected = [k for k, v in self.chkLevels.items() if v.isChecked()]
        self.filtered_words = []

        for w in self.all_words:
            if w["skip"] == "не повторять":
                continue

            for lvl in selected:
                if lvl == "C1+" and w["level"] == "C1":
                    self.filtered_words.append(w)
                    break
                if w["level"] == lvl:
                    self.filtered_words.append(w)
                    break

        self.show_word()

    # ---------------- WORD ----------------
    def show_word(self):
        if not self.filtered_words:
            self.current = None
            self.lblWord.setText("Нет слов")
            self.lblTranslation.setText("")
            return

        self.current = random.choice(self.filtered_words)

        self.lblTranslation.setText(self.current["translation"])
        self.update_view()
        self.lblResult.setText("")

    def update_view(self):
        if not self.current:
            return

        word = self.current["word"]

        if word.startswith("der "):
            article, color = "der", "#00f5ff"
        elif word.startswith("die "):
            article, color = "die", "#ff4dff"
        elif word.startswith("das "):
            article, color = "das", "#ffd166"
        else:
            article, color = "", "white"

        if self.chkNoColor.isChecked():
            color = "white"

        text = word if self.chkShowArticle.isChecked() else (
            word[len(article)+1:] if article else word
        )

        if self.chkHideTranslation.isChecked():
            self.lblTranslation.setText("")

        self.lblWord.setText(text)
        self.lblWord.setStyleSheet(f"color:{color};")

    # ---------------- CHECK ----------------
    def check(self, a):
        if not self.current:
            return

        word = self.current["word"]

        if word.startswith(a + " "):
            self.lblResult.setText("✔ правильно")
            self.lblResult.setStyleSheet("color:#00ff88;")
            self.show_word()
        else:
            self.session_errors += 1
            self.lblResult.setText("✖ неверно")
            self.lblResult.setStyleSheet("color:#ff3b3b;")
            self.lblErrors.setText(f"Ошибки: {self.session_errors}")

    # ---------------- RESET ----------------
    def reset_errors(self):
        self.session_errors = 0
        self.lblErrors.setText("Ошибки: 0")

    # ---------------- EDIT CURRENT WORD ----------------
    def open_editor(self):
        if not self.current:
            return

        dlg = QDialog(self)
        uic.loadUi("editor.ui", dlg)

        dlg.txtWord.setText(self.current["word"])
        dlg.txtTrans.setText(self.current["translation"])
        dlg.txtLevel.setText(self.current["level"])

        def save():
            word = dlg.txtWord.text().strip()
            tr = dlg.txtTrans.text().strip()
            lvl = dlg.txtLevel.text().strip()

            if not word or not tr:
                QMessageBox.warning(dlg, "Ошибка", "Заполни слово и перевод")
                return

            ws = self.current["sheet"]
            r = self.current["row"]

            ws[f"A{r}"] = word
            ws[f"B{r}"] = tr
            ws[f"C{r}"] = lvl

            self.wb.save(EXCEL_FILE)

            self.current["word"] = word
            self.current["translation"] = tr
            self.current["level"] = lvl

            self.lblTranslation.setText(tr)
            self.update_view()

            dlg.close()

        dlg.btnSave.clicked.connect(save)
        dlg.btnCancel.clicked.connect(dlg.close)

        dlg.exec()

    # ---------------- ADD WORD ----------------
    def add_word(self):
        dlg = QDialog(self)
        uic.loadUi("editor.ui", dlg)

        def save():
            word = dlg.txtWord.text().strip()
            tr = dlg.txtTrans.text().strip()
            lvl = dlg.txtLevel.text().strip()

            if not word or not tr:
                QMessageBox.warning(dlg, "Ошибка", "Заполни слово и перевод")
                return

            ws = self.wb["other"]
            r = ws.max_row + 1

            ws[f"A{r}"] = word
            ws[f"B{r}"] = tr
            ws[f"C{r}"] = lvl

            self.wb.save(EXCEL_FILE)
            dlg.close()

        dlg.btnSave.clicked.connect(save)
        dlg.btnCancel.clicked.connect(dlg.close)

        dlg.exec()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = App()
    w.show()
    sys.exit(app.exec())